# -*- coding: utf-8 -*-

from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtSql import QSqlDatabase
from PyQt5.QtSql import QSqlQueryModel, QSqlTableModel
from back_end import *
import logging
import datetime
import os
import threading
import openpyxl
from openpyxl.styles import Border, Side, PatternFill
import sys

# переменные списков найденных таблиц для вывода, которые будут изменены через global, для дальнейшего удаления в
# функции delete_report
list_table_for_delete_report = []

# список флажков для выбора репортов
list_check_box = []
# активатор авторизации для отображения флажков
authorization = 0

# получаем имя машины с которой был осуществлён вход в программу
uname = os.environ.get('USERNAME')
# настраиваем систему логирования
# дата LogFile из системы
date_log_file = datetime.datetime.now().strftime("%m %Y")
new_path_log_file = ''
# путь к папке где будет сохраняться LogFile
new_path_log_file = os.path.abspath(os.getcwd()) + '\\Log File\\'
# если папка Log File не создана,
if not os.path.exists(new_path_log_file):
    # то создаём эту папку
    os.makedirs(new_path_log_file)
# путь сохранения + имя LogFile
name_log_file = new_path_log_file + date_log_file + ' Log File.txt'
logging.basicConfig(level=logging.INFO,
                    handlers=[logging.FileHandler(filename=name_log_file, mode='a', encoding='utf-8')],
                    format='%(asctime)s [%(levelname)s] Пользователь: %(user)s - %(message)s', )
# дополняем базовый формат записи лог сообщения данными о пользователе
logger = logging.getLogger()
logger_with_user = logging.LoggerAdapter(logger, {'user': uname})

logger_with_user.info('Запуск программы')

# создаём приложение
app = QApplication(sys.argv)
# создаём окно приложения
window = QWidget()
# название приложения
window.setWindowTitle('Data Search Engine')
# задаём стиль приложения Fusion
app.setStyle('Fusion')
# размер окна приложения
window.setFixedSize(1521, 965)
# window.setFixedSize(1524, 872)

# устанавливаем favicon в окне приложения
icon = QIcon()
icon.addFile(u"icon.ico", QSize(), QIcon.Normal, QIcon.Off)
icon.addFile(u"icon.ico", QSize(), QIcon.Active, QIcon.On)

app.setWindowIcon(icon)

# создаём однострочное поле для ввода номер линии или чертежа
line_search = QLineEdit(window)
# устанавливаем положение окна ввода и его размеры в родительском окне
line_search.setGeometry(QRect(20, 20, 561, 41))
# присваиваем уникальное объектное имя однострочному полю для ввода
line_search.setObjectName(u"line_search")
# задаём параметры стиля и оформления окна ввода
font_line_search = QFont()
font_line_search.setFamily(u"Arial")
font_line_search.setPointSize(14)
font_line_search.setItalic(False)
# дополнительные параметры
line_search.setFont(font_line_search)
line_search.setMouseTracking(False)
line_search.setFocusPolicy(Qt.ClickFocus)
line_search.setContextMenuPolicy(Qt.NoContextMenu)
line_search.setAcceptDrops(True)
line_search.setStyleSheet(u"")
line_search.setLocale(QLocale(QLocale.English, QLocale.UnitedStates))
line_search.setEchoMode(QLineEdit.Normal)
line_search.setCursorPosition(0)
line_search.setCursorMoveStyle(Qt.LogicalMoveStyle)
line_search.setClearButtonEnabled(True)
line_search.setText('28278087')
line_search.setFocus()

# создаём кнопку "Поиск"
button_search = QPushButton('Поиск', window)
# устанавливаем положение и размер кнопки для поиска в родительском окне (window)
button_search.setGeometry(600, 20, 161, 41)
# присваиваем уникальное объектное имя кнопке "Поиск"
button_search.setObjectName(u"pushButton_search")
# задаём параметры стиля и оформления кнопки "Поиск"
font_button_search = QFont()
font_button_search.setFamily(u"Arial")
font_button_search.setPointSize(14)
button_search.setFont(font_button_search)
# дополнительные параметры
button_search.setFocusPolicy(Qt.ClickFocus)

# создаём кнопку "Закрыть" из программы
button_exit = QPushButton('Закрыть', window)
# устанавливаем положение и размер кнопки "Закрыть" для выхода из программы в родительском окне (window)
button_exit.setGeometry(QRect(1340, 904, 161, 41))
# присваиваем уникальное объектное имя кнопке "Закрыть"
button_exit.setObjectName(u"pushButton_exit")
# задаём параметры стиля и оформления кнопки "Закрыть"
font_button_exit = QFont()
font_button_exit.setFamily(u"Arial")
font_button_exit.setPointSize(14)
button_exit.setFont(font_button_exit)
# дополнительные параметры
button_exit.setFocusPolicy(Qt.ClickFocus)
# Закрытие программы при нажатии на кнопку "Закрыть"
button_exit.clicked.connect(qApp.exit)

# создаём однострочное поле для ввода логина
line_login = QLineEdit(window)
# присваиваем уникальное объектное имя полю для ввода логина
line_login.setObjectName(u"line_login")
# устанавливаем положение и размер поля для ввода логина в родительском окне (window)
line_login.setGeometry(QRect(1270, 20, 111, 31))
# задаём параметры стиля и оформления поля для ввода логина
font_line_login = QFont()
font_line_login.setFamily(u"Arial")
font_line_login.setPointSize(11)
font_line_login.setItalic(True)
line_login.setFont(font_line_login)
# дополнительные параметры
line_login.setEchoMode(QLineEdit.Normal)
# устанавливаем исчезающий текст
line_login.setPlaceholderText('login')
line_login.setText('admin')

# создаём однострочное поле для ввода пароля
line_password = QLineEdit(window)
# присваиваем уникальное объектное имя полю для ввода пароля
line_password.setObjectName(u"line_password")
# устанавливаем положение и размер поля для ввода пароля в родительском окне (window)
line_password.setGeometry(QRect(1270, 60, 111, 31))
# задаём параметры стиля и оформления поля для ввода пароля
font_line_password = QFont()
font_line_password.setFamily(u"Arial")
font_line_password.setPointSize(11)
font_line_password.setItalic(True)
line_password.setFont(font_line_password)
# дополнительные параметры
line_password.setEchoMode(QLineEdit.Password)
# устанавливаем исчезающий текст
line_password.setPlaceholderText('password')
line_password.setText('admin')

# устанавливаем надпись "Логин"
label_login = QLabel('Логин', window)
# присваиваем уникальное объектное имя надписи "Логин"
label_login.setObjectName(u"label_login")
# устанавливаем положение и размер поля для надписи "Логин" в родительском окне (window)
label_login.setGeometry(QRect(1200, 30, 61, 21))
# задаём параметры стиля и оформления поля для надписи "Логин"
font_label_login = QFont()
font_label_login.setFamily(u"Arial")
font_label_login.setPointSize(12)
font_label_login.setItalic(True)
label_login.setFont(font_label_login)

# устанавливаем надпись "Пароль"
label_password = QLabel('Пароль', window)
# присваиваем уникальное объектное имя надписи "Пароль"
label_password.setObjectName(u"label_password")
# устанавливаем положение и размер поля для надписи "Пароль" в родительском окне (window)
label_password.setGeometry(QRect(1190, 70, 81, 21))
# задаём параметры стиля и оформления поля для надписи "Пароль"
font_label_password = QFont()
font_label_password.setFamily(u"Arial")
font_label_password.setPointSize(12)
font_label_password.setItalic(True)
# скрываем введённые с клавиатуры символы при вводе в поле "Пароль"
label_password.setFont(font_label_password)

# создаём кнопку печати
button_print = QPushButton('Печать', window)
# устанавливаем положение и размер кнопки печати в родительском окне (window)
button_print.setGeometry(QRect(20, 155, 161, 41))
# присваиваем уникальное объектное имя кнопке "Печать"
button_print.setObjectName(u"pushButton_print")
# задаём параметры стиля и оформления кнопки печати
font_button_print = QFont()
font_button_print.setFamily(u"Arial")
font_button_print.setPointSize(14)
button_print.setFont(font_button_print)
# дополнительные параметры
button_print.setFocusPolicy(Qt.ClickFocus)

# создаём кнопку "Войти"
button_log_in = QPushButton('Войти', window)
# устанавливаем положение и размер кнопки "Войти" в родительском окне (window)
button_log_in.setGeometry(QRect(1390, 20, 111, 31))
# присваиваем уникальное объектное имя кнопке "Войти"
button_log_in.setObjectName(u"pushButton_enter")
# задаём параметры стиля и оформления кнопки "Войти"
font_button_log_in = QFont()
font_button_log_in.setFamily(u"Arial")
font_button_log_in.setPointSize(14)
button_log_in.setFont(font_button_log_in)
# дополнительные параметры
button_log_in.setFocusPolicy(Qt.ClickFocus)

# создаём кнопку "Выйти"
button_log_out = QPushButton('Выйти', window)
# устанавливаем положение и размер кнопки "Выйти" в родительском окне (window)
button_log_out.setGeometry(QRect(1390, 60, 111, 31))
# присваиваем уникальное объектное имя кнопке "Выйти"
button_log_out.setObjectName(u"pushButton_out")
# задаём параметры стиля и оформления кнопки "Выйти"
font_button_log_out = QFont()
font_button_log_out.setFamily(u"Arial")
font_button_log_out.setPointSize(14)
button_log_out.setFont(font_button_log_out)
# дополнительные параметры
button_log_out.setFocusPolicy(Qt.ClickFocus)
# делаем неактивной кнопку "Выйти" до авторизации
button_log_out.setDisabled(True)

# создаём кнопку "Добавить"
button_add = QPushButton('Добавить', window)
# устанавливаем положение и размер кнопки "Добавить" в родительском окне (window)
button_add.setGeometry(QRect(200, 155, 161, 41))
# присваиваем уникальное объектное имя кнопке "Добавить"
button_add.setObjectName(u"pushButton_add")
# задаём параметры стиля и оформления кнопки "Добавить"
font_button_add = QFont()
font_button_add.setFamily(u"Arial")
font_button_add.setPointSize(14)
button_add.setFont(font_button_add)
# дополнительные параметры
button_add.setFocusPolicy(Qt.ClickFocus)
# делаем неактивной кнопку "Добавить" до авторизации
button_add.setDisabled(True)

# создаём кнопку "Удалить"
button_delete = QPushButton('Удалить', window)
# устанавливаем положение и размер кнопки "Удалить" в родительском окне (window)
button_delete.setGeometry(QRect(20, 904, 171, 41))
# присваиваем уникальное объектное имя кнопке "Удалить"
button_delete.setObjectName(u"pushButton_delete")
# задаём параметры стиля и оформления кнопки "Удалить"
font_button_delete = QFont()
font_button_delete.setFamily(u"Arial")
font_button_delete.setPointSize(14)
button_delete.setFont(font_button_delete)
# дополнительные параметры
button_delete.setFocusPolicy(Qt.ClickFocus)
# делаем неактивной кнопку "Удалить" до авторизации
button_delete.setDisabled(True)

# создаём кнопку "Создать отчёт"
button_create_report = QPushButton('Создать отчёт', window)
# устанавливаем положение и размер кнопки "Создать отчёт" в родительском окне (window)
button_create_report.setGeometry(QRect(380, 155, 201, 41))
# присваиваем уникальное объектное имя кнопке "Сформировать отчёт"
button_create_report.setObjectName(u"pushButton_create_report")
# задаём параметры стиля и оформления кнопки "Сформировать отчёт"
font_button_create_report = QFont()
font_button_create_report.setFamily(u"Arial")
font_button_create_report.setPointSize(14)
button_create_report.setFont(font_button_create_report)
# дополнительные параметры
button_create_report.setFocusPolicy(Qt.ClickFocus)

# создаём кнопку "Сводные данные"
button_statistic_master = QPushButton('Сводные данные', window)
# устанавливаем положение и размер кнопки "Сводные данные" в родительском окне (window)
button_statistic_master.setGeometry(QRect(659, 904, 201, 41))
# присваиваем уникальное объектное имя кнопке "Сводные данные"
button_statistic_master.setObjectName(u"pushButton_statistic_master")
# задаём параметры стиля и оформления кнопки "Сводные данные"
font_button_statistic_master = QFont()
font_button_statistic_master.setFamily(u"Arial")
font_button_statistic_master.setPointSize(14)
button_statistic_master.setFont(font_button_statistic_master)
# дополнительные параметры
button_statistic_master.setFocusPolicy(Qt.ClickFocus)
# делаем неактивной кнопку "Сводные данные" до авторизации
button_statistic_master.setDisabled(True)

# вставляем картинку YKR
label_ykr = QLabel(window)
label_ykr.setObjectName(u"Rutledge")
label_ykr.setGeometry(QRect(790, 10, 111, 121))
label_ykr.setPixmap(QPixmap(u"logo_ykr.png"))

# вставляем картинку NCA
label_nca = QLabel(window)
label_nca.setObjectName(u"Rutledge")
label_nca.setGeometry(QRect(920, 10, 111, 121))
label_nca.setPixmap(QPixmap(u"logo_nca.png"))

# вставляем картинку NCOC
label_ncoc = QLabel(window)
label_ncoc.setObjectName(u"Rutledge")
label_ncoc.setGeometry(QRect(1050, 13, 111, 115))
label_ncoc.setPixmap(QPixmap(u"logo_ncoc.png"))

# общая область с боковой полосой прокрутки
scroll_area = QScrollArea(window)
scroll_area.setObjectName(u'Scroll_Area')
# полоса прокрутки появляется, только если таблицы больше самой области прокрутки
scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
# задаём размер области с полосой прокрутки
scroll_area.setGeometry(20, 215, 1481, 670)

# создаём группу из радио-кнопок 'ON', 'OF', 'OS'
groupBox_location = QGroupBox(window)
groupBox_location.setObjectName(u"groupBox_radio")
# устанавливаем размер группы радио-кнопок
groupBox_location.setGeometry(QRect(20, 80, 161, 56))
# устанавливаем название группы радио-кнопок
groupBox_location.setTitle('Локация')
groupBox_location.setStyleSheet('''QGroupBox {border: 0.5px solid grey;};
                                   QGroupBox:title{
                                   subcontrol-origin: margin;
                                   subcontrol-position: top center;
                                   padding: 0 3px 0 3px;
                                }''')

# создаём радио-кнопку локации 'ON'
radioButton_on = QRadioButton(groupBox_location)
radioButton_on.setObjectName(u"radioButton")
# устанавливаем положение внутри группы
radioButton_on.setGeometry(QRect(10, 25, 42, 20))
# указываем текст радио-кнопки
radioButton_on.setText('ON')
# делаем радио-кнопку 'ON' активной по умолчанию
radioButton_on.setChecked(True)

# создаём радио-кнопку локации 'OF'
radioButton_of = QRadioButton(groupBox_location)
radioButton_of.setObjectName(u"radioButton_2")
# устанавливаем положение внутри группы
radioButton_of.setGeometry(QRect(60, 25, 42, 20))
# указываем текст радио-кнопки
radioButton_of.setText('OF')

# создаём радио-кнопку локации 'OS'
radioButton_os = QRadioButton(groupBox_location)
radioButton_os.setObjectName(u"radioButton_3")
# устанавливаем положение внутри группы
radioButton_os.setGeometry(QRect(110, 25, 42, 20))
# указываем текст радио-кнопки
radioButton_os.setText('OS')

# создаём группу из чек-боксов методов контроля
groupBox_ndt = QGroupBox(window)
groupBox_ndt.setObjectName(u"groupBox_ndt")
groupBox_ndt.setGeometry(QRect(200, 80, 161, 56))
# устанавливаем название группы чек-боксов
groupBox_ndt.setTitle('Метод контроля')
groupBox_ndt.setStyleSheet('''QGroupBox {border: 0.5px solid grey;};
                              QGroupBox:title{
                              subcontrol-origin: margin;
                              subcontrol-position: top center;
                              padding: 0 3px 0 3px;
                           }''')

# создаём чек-бокс метода контроля 'UTT'
checkBox_utt = QCheckBox(groupBox_ndt)
checkBox_utt.setObjectName(u"checkBox_utt")
# устанавливаем положение внутри группы
checkBox_utt.setGeometry(QRect(10, 25, 61, 20))
# указываем текст чек-бокса
checkBox_utt.setText('UTT')
# делаем чек-бокс 'UTT' активным по умолчанию
checkBox_utt.setChecked(True)

# создаём чек-бокс метода контроля 'PAUT'
checkBox_paut = QCheckBox(groupBox_ndt)
checkBox_paut.setObjectName(u"checkBox_paut")
# устанавливаем положение внутри группы
checkBox_paut.setGeometry(QRect(80, 25, 61, 20))
# указываем текст чек-бокса
checkBox_paut.setText('PAUT')

# создаём группу из чек-боксов годов
groupBox_year = QGroupBox(window)
groupBox_year.setObjectName(u"groupBox_year")
# устанавливаем размер группы радио-кнопок
groupBox_year.setGeometry(QRect(380, 80, 381, 56))
# устанавливаем название группы чек-боксов
groupBox_year.setTitle('Год контроля')
groupBox_year.setStyleSheet('''QGroupBox {border: 0.5px solid grey;};
                               QGroupBox:title{
                               subcontrol-origin: margin;
                               subcontrol-position: top center;
                               padding: 0 3px 0 3px;
                            }''')

# создаём чек-бокс года '2023'
checkBox_2023 = QCheckBox(groupBox_year)
checkBox_2023.setObjectName(u"checkBox_2023")
# устанавливаем положение внутри группы
checkBox_2023.setGeometry(QRect(10, 25, 61, 20))
# указываем текст чек-бокса
checkBox_2023.setText('2023')

# создаём чек-бокс года '2022'
checkBox_2022 = QCheckBox(groupBox_year)
checkBox_2022.setObjectName(u"checkBox_2022")
# устанавливаем положение внутри группы
checkBox_2022.setGeometry(QRect(80, 25, 61, 20))
# указываем текст чек-бокса
checkBox_2022.setText('2022')
# делаем чек-бокс '2022' активным по умолчанию
checkBox_2022.setChecked(True)

# создаём чек-бокс года '2021'
checkBox_2021 = QCheckBox(groupBox_year)
checkBox_2021.setObjectName(u"checkBox_2021")
# устанавливаем положение внутри группы
checkBox_2021.setGeometry(QRect(150, 25, 61, 20))
# указываем текст чек-бокса
checkBox_2021.setText('2021')

# создаём чек-бокс года '2020'
checkBox_2020 = QCheckBox(groupBox_year)
checkBox_2020.setObjectName(u"checkBox_2020")
# устанавливаем положение внутри группы
checkBox_2020.setGeometry(QRect(220, 25, 61, 20))
# указываем текст чек-бокса
checkBox_2020.setText('2020')

# создаём чек-бокс года '2019'
checkBox_2019 = QCheckBox(groupBox_year)
checkBox_2019.setObjectName(u"checkBox_2019")
# устанавливаем положение внутри группы
checkBox_2019.setGeometry(QRect(290, 25, 61, 20))
# указываем текст чек-бокса
checkBox_2019.setText('2019')


# нажатие на кнопку "Добавить"
def add_tables():
    # выбираем один или несколько файлов с расширением docx
    name_dir = QFileDialog.getOpenFileNames(None, 'Выбрать папку', r'C:\Users\asus\Documents\NDT YKR\NDT UTT\REPORTS 2022\UTT\ON',
                                            "docx(*.docx)")
    # name_dir = QFileDialog.getOpenFileNames(None, 'Выбрать папку', '/home', "docx(*.docx)")
    # если выбран файл репорт для загрузки, то
    if name_dir[0]:
        # анимация загрузки
        gif_loading()
        # создаём отдельный поток для выполнения добавления новых репортов
        thr_add_table = threading.Thread(target=add_table, args=(name_dir,))
        thr_add_table.start()
        logger_with_user.info('Добавление новых репортов в базу данных')


# список для вывода в файл Excel для печати
data_for_print = []
# список названий столбцов для вывода в файл Excel для печати
name_column_for_print = []
# список найденных моделей с данными из базы данных
list_sqm = []
# список названий листов Excel при печати по названиям кнопок при поиске репортов
list_name_sheet_for_print = []
# список минимальных значений
list_min_thickness = []
# список минимальных значений толщин в каждом столбце
list_name_sheet_for_print = []
# индекс столбца с номинальной толщиной
index_nom_thickness_name_column = ''


# нажатие на кнопку "Поиск"
def search():
    global data_for_print, name_column_for_print, list_sqm, list_min_thickness, list_name_sheet_for_print
    # обнуляем список минимальных значений толщин в каждом столбце
    list_name_sheet_for_print = []
    # обнуляем список минимальных значений
    list_min_thickness = []
    # обнуляем список для вывода в файл Excel для печати
    data_for_print = []
    # обнуляем список названий столбцов для вывода в файл Excel для печати
    name_column_for_print = []
    # обнуляем список найденных моделей с данными из базы данных
    list_sqm = []
    # список фильтров годов для поиска
    year_db = []
    # список фильтров методов контроля для поиска
    ndt_db = []
    # список баз данных для поиска по выбранным фильтрам
    db_for_search = []
    # список баз данных для поиска только с годами, для дальнейшего прибавления методов контроля
    db_only_ndt_for_search = []
    # активатор, если вызвана статистика
    global check_statistic_master
    check_statistic_master = 0

    # проверяем наличие областей tableView для вывода данных
    # если есть, то закрываем их, чтобы не наслаивались
    if window.findChildren(QTableView):
        open_tableview = window.findChildren(QTableView)
        for i in open_tableview:
            i.hide()
    if line_search.text():
        # удаляем обозначения дюймов "
        if re.findall(r'\'\'|"|”', line_search.text()):
            line_for_search = re.sub(r'"|\'\'', '', line_search.text()).upper()
        else:
            # получаем текст из поля для ввода и приводим его в верхний регистр
            line_for_search = line_search.text().upper()

        # проверяем какие фильтры локации установлены для поиска
        if radioButton_on.isChecked():
            location_db = 'ON'
        if radioButton_of.isChecked():
            location_db = 'OF'
        if radioButton_os.isChecked():
            location_db = 'OS'
        # проверяем какие фильтры годов установлены для поиска
        if checkBox_2019.isChecked():
            year_db.append('19')
        if checkBox_2020.isChecked():
            year_db.append('20')
        if checkBox_2021.isChecked():
            year_db.append('21')
        if checkBox_2022.isChecked():
            year_db.append('22')
        if checkBox_2023.isChecked():
            year_db.append('23')
        # если ни один год не выбран, то выдаём сообщение об этом и останавливаем выполнение поиска
        if not year_db:
            return QMessageBox.information(
                window,
                'Внимание!',
                'Вы не выбрали год(а) для поиска!!!',
                buttons=QMessageBox.Ok
            )
        # проверяем какие фильтры методов контроля установлены для поиска
        if checkBox_utt.isChecked():
            ndt_db.append('UTT')
        if checkBox_paut.isChecked():
            ndt_db.append('PAUT')
        # если ни один метод контроля не выбран, то выдаём сообщение об этом и останавливаем выполнение поиска
        if not ndt_db:
            return QMessageBox.information(
                window,
                'Внимание!',
                'Вы не выбрали ни один метод контроля для поиска!!!',
                buttons=QMessageBox.Ok
            )

        # собираем название баз данных для поиска по выбранным фильтрам
        # вначале собираем список баз данных из выбранных методов контроля
        for ndt_db_for_search in ndt_db:
            db_only_ndt_for_search.append('reports_db_' + location_db + '_' + ndt_db_for_search)
        # затем к списку баз данных с методами контроля добавляем выбранные года
        for year_db_for_search in year_db:
            for i in db_only_ndt_for_search:
                db_for_search.append(i + '_' + year_db_for_search + '.sqlite')

        # перебираем базы данных по выбранным фильтрам для поиска данных
        for db in db_for_search:
            # создаём соединение с базой данной
            con = QSqlDatabase.addDatabase('QSQLITE')
            # передаём имя базы данных для открытия
            con.setDatabaseName((r'C:\Users\asus\PycharmProjects\YKR\YKR\DB\\' + db))
            # con.setDatabaseName(r'C:\Users\asus\PycharmProjects\YKR\YKR\reports_db.sqlite')
            # если соединение не установлено, то сообщение об ошибке и выход
            if not con.open():
                QMessageBox.critical(
                    None,
                    'App name Error',
                    'Error to connect to the database')
                logger_with_user.error('Отсутствует соединение с базой данных')
                sys.exit()
            else:
                # список таблиц в которой есть искомая линия
                table_for_search_line = []
                # список таблиц в которой есть искомый чертёж
                table_for_search_drawing = []
                # список таблиц в которой есть искомый номер репорта
                table_for_search_report = []
                # список таблиц в которой есть искомый номер work order
                table_for_search_wo = []
                # список репортов из sqlite_master где есть номер work order
                reports_for_search_wo = []
                # проверяем наличие областей tableView для вывода данных
                # если есть, то закрываем их, чтобы не наслаивались
                if window.findChildren(QTableView):
                    open_tableview = window.findChildren(QTableView)
                    for i in open_tableview:
                        i.hide()

                # если в поле для поиска указан номер репорта
                if line_for_search[:6] == '04-YKR':
                    # перебираем sqlite_master в поиске репорта
                    # подключаемся в базе данных
                    conn = sqlite3.connect(r'C:\Users\asus\PycharmProjects\YKR\YKR\DB\\' + db)
                    # conn = sqlite3.connect('reports_db.sqlite')
                    # conn.isolation_level = None
                    cur = conn.cursor()
                    # меняем '-' на '_'
                    line_for_search_report = re.sub('-', '_', line_for_search)
                    # если нашли номер репорта в sqlite_master
                    if cur.execute('SELECT tbl_name FROM sqlite_master WHERE name LIKE "%{}"'.format(line_for_search_report)):
                        table_for_search_report.append(
                            cur.execute('SELECT tbl_name FROM sqlite_master WHERE name LIKE "%{}"'.format(
                                line_for_search_report)).fetchall())
                    cur.close()

                # если в поле для поиска указан номер work order
                elif line_for_search.isdigit():
                    # перебираем master в поиске work order
                    # подключаемся в базе данных
                    conn = sqlite3.connect(r'C:\Users\asus\PycharmProjects\YKR\YKR\DB\\' + db)
                    # conn = sqlite3.connect('reports_db.sqlite')
                    # conn.isolation_level = None
                    cur = conn.cursor()
                    # если нашли work order в master
                    if cur.execute('SELECT report_number FROM master WHERE work_order="{}"'.format(line_for_search)):
                        reports_for_search_wo.append(
                            cur.execute('SELECT report_number FROM master WHERE work_order="{}"'.format(line_for_search)).fetchall())
                        # перебираем найденные номера репортов
                        for ii in reports_for_search_wo[0]:
                            # добавляем в список если нашли номера таблиц
                            table_for_search_wo.append(
                                cur.execute('SELECT tbl_name FROM sqlite_master WHERE name LIKE "%{}"'.format(ii[0])).fetchall())
                    # закрываем соединение
                    cur.close()

                else:
                    # перебираем таблицы, которые попали в базу данных после очистки
                    for i in con.tables():
                        # подключаемся в базе данных
                        # conn = sqlite3.connect(r'C:\Users\asus\PycharmProjects\YKR\YKR\DB\\' + db)
                        conn = sqlite3.connect(r'C:\Users\asus\PycharmProjects\YKR\YKR\DB\\' + db)
                        # conn = sqlite3.connect('reports_db.sqlite')
                        # conn.isolation_level = None
                        cur = conn.cursor()
                        # перебираем список названий столбцов в таблице
                        for k in cur.execute('SELECT * FROM {}'.format(i)).description:
                            # если 'Line' есть в названии столбца
                            if 'Line' in k:
                                # и если искомая линия есть в таблице, то добавляем имя таблицы в список table_for_search_line
                                if cur.execute('SELECT Line FROM {} WHERE Line LIKE "%{}%"'.format(i, line_for_search)).fetchall():
                                    table_for_search_line.append(i)
                            # если 'Drawing' есть в названии столбца
                            if 'Drawing' in k:
                                # и если искомый чертёж есть в таблице, то добавляем имя таблицы в список
                                # table_for_search_drawing
                                if cur.execute(
                                        'SELECT Drawing FROM {} WHERE Drawing LIKE "%{}%"'.format(i, line_for_search)).fetchall():
                                    table_for_search_drawing.append(i)
                        cur.close()





                # если для поиска указан не номер репорта или work order
                def depthCount(x, depth=0):
                    if not x or not isinstance(x, list):
                        return depth
                    return max(depthCount(x[0], depth + 1),
                               depthCount(x[1:], depth))

                # если для поиска указан не номер репорта
                if depthCount(table_for_search_report) == 0:
                    table_for_search_report = [[]]

                # если для поиска указан не номер work order
                if depthCount(table_for_search_wo) == 0:
                    table_for_search_wo = [[]]

                # если найден номер линии или номер чертежа, или номер репорта, или номер work order, то показываем область
                # для таблицы с найденными данными
                if table_for_search_line or table_for_search_drawing or table_for_search_report or table_for_search_wo:
                    if len(table_for_search_line) + len(table_for_search_drawing) + len(
                            table_for_search_report[0]) + len(table_for_search_wo[0]) != 0:
                        # считаем количество найденных таблиц для вывода нужного количества tableView
                        count_table_view = len(table_for_search_line) + len(table_for_search_drawing) + len(
                            table_for_search_report[0]) + len(table_for_search_wo[0])
                        # список названий таблицы для переменной при создании tableView
                        table_view = ['one_table', 'two_table', 'three_table', 'four_table', 'five_table', 'six_table',
                                      'seven_table', 'eight_table', 'nine_table', 'ten_table', 'eleven_table',
                                      'twelve_table', 'thirteen_table', 'fourteen_table', 'fifteen_table', 'sixteen_table',
                                      'seventeen_table', 'eighteen_table', 'nineteen_table', 'twenty_table',
                                      'twenty_one_table', 'twenty_two_table', 'twenty_three_table', 'twenty_four_table',
                                      'twenty_five_table', 'twenty_six_table', 'twenty_seven_table', 'twenty_eight_table',
                                      'twenty_nine_table', 'thirty_table', 'thirty_one_table', 'thirty_two_table',
                                      'thirty_three_table', 'thirty_four_table', 'thirty_five_table', 'thirty_six_table',
                                      'thirty_seven_table', 'thirty_eight_table', 'thirty_nine_table', 'forty_table',
                                      'forty_one_table', 'forty_two_table', 'forty_three_table', 'forty_four_table',
                                      'forty_five_table', 'forty_six_table', 'forty_seven_table', 'forty_eight_table',
                                      'forty_nine_table', 'fifty_table']
                        # список check box для переменной check_box (флажок выбора выведенного репорта)
                        check_box = ['one_check_box', 'two_check_box', 'three_check_box', 'four_check_box',
                                     'five_check_box', 'six_check_box', 'seven_check_box', 'eight_check_box',
                                     'nine_check_box', 'ten_check_box', 'eleven_check_box', 'twelve_check_box',
                                     'thirteen_check_box', 'fourteen_check_box', 'fifteen_check_box',
                                     'sixteen_check_box', 'seventeen_check_box', 'eighteen_check_box', 'nineteen_check_box',
                                     'twenty_check_box', 'twenty_one_check_box', 'twenty_two_check_box',
                                     'twenty_three_check_box', 'twenty_four_check_box', 'twenty_five_check_box',
                                     'twenty_six_check_box', 'twenty_seven_check_box', 'twenty_eight_check_box',
                                     'twenty_nine_check_box', 'thirty_check_box', 'thirty_one_check_box',
                                     'thirty_two_check_box', 'thirty_three_check_box', 'thirty_four_check_box',
                                     'thirty_five_check_box', 'thirty_six_check_box', 'thirty_seven_check_box',
                                     'thirty_eight_check_box', 'thirty_nine_check_box', 'forty_check_box',
                                     'forty_one_check_box', 'forty_two_check_box', 'forty_three_check_box',
                                     'forty_four_check_box', 'forty_five_check_box', 'forty_six_check_box',
                                     'forty_seven_check_box', 'forty_eight_check_box', 'forty_nine_check_box',
                                     'fifty_check_box']
                        # frame в который будут вставляться, таблицы чтобы при большом количестве таблиц появлялась полоса прокрутки
                        frame_for_table = QFrame()

                        # список количества строк в каждой найденной таблице
                        count_row_table_view = []
                        global list_table_for_delete_report
                        list_table_for_delete_report = []
                        if table_for_search_line:
                            # подключаемся в базе данных
                            conn = sqlite3.connect(r'C:\Users\asus\PycharmProjects\YKR\YKR\DB\\' + db)
                            # conn = sqlite3.connect('reports_db.sqlite')
                            # подключаемся в базе данных
                            cur = conn.cursor()
                            # изменяем первоначальную переменную на список таблиц для дальнейшего удаления
                            # list_table_for_delete_report = table_for_search_line
                            list_table_for_delete_report.append(table_for_search_line)

                            for i in table_for_search_line:
                                # количество строк в одной найденной таблице count_row_table_view[0][0]
                                count_row_table = cur.execute(
                                    'SELECT COUNT(*) FROM {} WHERE Line LIKE "%{}%"'.format(i, line_for_search)).fetchall()
                                count_row_table_view.append(count_row_table[0][0])
                            cur.close()
                        if table_for_search_drawing:
                            # подключаемся в базе данных
                            conn = sqlite3.connect(r'C:\Users\asus\PycharmProjects\YKR\YKR\DB\\' + db)
                            # conn = sqlite3.connect('reports_db.sqlite')
                            # подключаемся в базе данных
                            cur = conn.cursor()
                            # изменяем первоначальную переменную на список таблиц для дальнейшего удаления
                            # list_table_for_delete_report = table_for_search_drawing
                            list_table_for_delete_report.append(table_for_search_drawing)
                            for i in table_for_search_drawing:
                                # количество строк в одной найденной таблице count_row_table[0][0]
                                count_row_table = cur.execute(
                                    'SELECT COUNT(*) FROM {} WHERE Drawing LIKE "%{}%"'.format(i, line_for_search)).fetchall()
                                count_row_table_view.append(count_row_table[0][0])
                            cur.close()
                        if table_for_search_report:
                            # подключаемся в базе данных
                            conn = sqlite3.connect(r'C:\Users\asus\PycharmProjects\YKR\YKR\DB\\' + db)
                            # conn = sqlite3.connect('reports_db.sqlite')
                            # подключаемся в базе данных
                            cur = conn.cursor()
                            # изменяем первоначальную переменную на список таблиц для дальнейшего удаления
                            if table_for_search_report[0]:
                                # list_table_for_delete_report = table_for_search_report[0][0][0]
                                list_table_for_delete_report.append(table_for_search_report[0])
                            for i in table_for_search_report[0]:
                                count_row_table = cur.execute('SELECT COUNT(*) FROM {}'.format(i[0])).fetchall()
                                count_row_table_view.append(count_row_table[0][0])
                            cur.close()
                        if table_for_search_wo:
                            # подключаемся в базе данных
                            conn = sqlite3.connect(r'C:\Users\asus\PycharmProjects\YKR\YKR\DB\\' + db)
                            # conn = sqlite3.connect('reports_db.sqlite')
                            # подключаемся в базе данных
                            cur = conn.cursor()
                            # изменяем первоначальную переменную на список таблиц для дальнейшего удаления
                            if table_for_search_wo[0]:
                                # list_table_for_delete_report = table_for_search_wo[0][0][0]
                                list_table_for_delete_report.append(table_for_search_wo[0])
                            for i in table_for_search_wo[0]:
                                count_row_table = cur.execute('SELECT COUNT(*) FROM {}'.format(i[0])).fetchall()
                                count_row_table_view.append(count_row_table[0][0])
                            cur.close()
                        # закрываем соединение
                        # cur.close()
                        # общее количество строк в найденных таблицах для длины frame
                        sum_row_table = 0
                        for i in count_row_table_view:
                            sum_row_table += i
                        # высота одной строки
                        one_row = 25
                        # высота фрейма = общее количество строк в найденных таблицах * высоту одной строки +
                        # + количество таблиц * 2 (кнопка номера репорта и строка названий столбцов) * 20 (высота одной
                        # строки) + 20 (высота первой строки с номером первого репорта) + количество таблиц * 20 (расстояние
                        # между таблицами в открытом виде
                        w = sum_row_table * one_row + len(count_row_table_view) * 2 * 20 + 20 + len(count_row_table_view) * 20
                        # помещаем frame в область с полосой прокрутки
                        scroll_area.setWidget(frame_for_table)
                        # задаём размер frame
                        frame_for_table.setGeometry(0, 0, 1460, w)
                        frame_for_table.show()
                        # начальная координата y1 - первой кнопки с номером репорта первой, y2 - первой таблицы
                        y1 = 0
                        # список всех таблиц, номеров репортов, номеров флажков и высоты каждой таблицы
                        list_table_view = []
                        list_button_for_table = []
                        global list_check_box
                        list_check_box = []
                        list_height_table_view = []
                        # вытягиваем данные из найденных таблиц, формируем таблицу, кнопку названия
                        for i in range(count_table_view):
                            # список названий столбцов отсчитывая от 'Nominal_thickness
                            name_column_for_min_thickness = []

                            # определяем глубину вложенности списка заданного для поиска репорта
                            if depthCount(table_for_search_report) == 1:
                                table_for_search_report = []

                            # определяем глубину вложенности списка заданного для поиска work order
                            if depthCount(table_for_search_wo) == 1:
                                table_for_search_wo = []
                            # подключаемся в базе данных
                            conn = sqlite3.connect(r'C:\Users\asus\PycharmProjects\YKR\YKR\DB\\' + db)
                            # conn = sqlite3.connect('reports_db.sqlite')
                            # conn.isolation_level = None
                            cur = conn.cursor()
                            # перебираем таблицы и извлекаем данные
                            if table_for_search_line:
                                print(table_for_search_line)
                                print(count_table_view)
                                print(i)
                                reader = cur.execute("SELECT * FROM {}".format(table_for_search_line[i]))
                            elif table_for_search_drawing:
                                reader = cur.execute("SELECT * FROM {}".format(table_for_search_drawing[i]))
                            elif table_for_search_report:
                                reader = cur.execute("SELECT * FROM {}".format(table_for_search_report[0][i][0]))
                            elif table_for_search_wo:
                                reader = cur.execute("SELECT * FROM {}".format(table_for_search_wo[0][i][0]))
                            # получаем список названий столбцов
                            name_column = [x[0] for x in reader.description]
                            cur.close()
                            # находим минимальное значение в выводимых данных
                            # список минимальных значений толщин в каждом столбце
                            list_min_thickness_column = []
                            # определяем индекс Nominal_thickness, что искать минимальное значение после него
                            global index_nom_thickness_name_column
                            index_nom_thickness_name_column = name_column.index('Nominal_thickness')
                            # формируем названия столбцов для поиска минимальной толщины без учета ненужных столбцов
                            for ij in range(index_nom_thickness_name_column + 1, len(name_column)):
                                name_column_for_min_thickness.append(name_column[ij])
                            # если название столбца не...
                            # for ii in name_column:
                            for ii in name_column_for_min_thickness:
                                if ii == 'Line' or ii == 'Item_description' or ii == 'Section' or ii == 'Location' \
                                        or ii == 'Remark' or ii == 'Size' or ii == 'Nominal_thickness' or ii == 'Diameter' \
                                        or ii == 'Drawing' or ii == 'P_ID' or ii == 'Date' or ii == 'Distance' \
                                        or ii == 'Result' or ii == 'S_N':
                                    continue
                                # то получаем все значения в столбце с измеренными толщинами
                                else:
                                    # подключаемся в базе данных
                                    conn = sqlite3.connect(r'C:\Users\asus\PycharmProjects\YKR\YKR\DB\\' + db)
                                    # conn = sqlite3.connect('reports_db.sqlite')
                                    # conn.isolation_level = None
                                    cur = conn.cursor()
                                    # список только вещественных чисел значений толщин в столбце
                                    list_thickness_column = []
                                    # определяем глубину вложенности списка заданного для поиска репорта
                                    if depthCount(table_for_search_report) == 1:
                                        table_for_search_report = []
                                    # определяем глубину вложенности списка заданного для поиска work order
                                    if depthCount(table_for_search_wo) == 1:
                                        table_for_search_wo = []
                                    # переменная всех значений толщин в столбце при поиске по номеру линии
                                    if table_for_search_line:
                                        thickness_column = cur.execute('SELECT {} from {}'.format(ii, table_for_search_line[i])).fetchall()
                                        # выбираем только вещественные значения
                                        for iii in thickness_column:
                                            # проверка если в столбце нет значений, то дальше, иначе...
                                            if not iii[0]:
                                                continue
                                            else:
                                                try:
                                                    if float(iii[0]):
                                                        list_thickness_column.append(float(iii[0]))
                                                except ValueError:
                                                    continue
                                        # если в столбце нет значений, то дальше, иначе...
                                        if not list_thickness_column:
                                            continue
                                        else:
                                            # минимальное значение толщины в столбце
                                            min_thickness_column = min(list_thickness_column)
                                            # добавляем это значение в список минимальных значений столбцов
                                            list_min_thickness_column.append(min_thickness_column)
                                    # переменная всех значений толщин в столбце при поиске по номеру чертежа
                                    if table_for_search_drawing:
                                        thickness_column = cur.execute('SELECT {} from {}'.format(ii, table_for_search_drawing[i])).fetchall()
                                        # выбираем только вещественные значения
                                        for iii in thickness_column:
                                            # проверка если в столбце нет значений, то дальше, иначе...
                                            if not iii[0]:
                                                continue
                                            else:
                                                try:
                                                    if float(iii[0]):
                                                        list_thickness_column.append(float(iii[0]))
                                                except ValueError:
                                                    continue
                                        # если в столбце нет значений, то дальше, иначе...
                                        if not list_thickness_column:
                                            continue
                                        else:
                                            # минимальное значение толщины в столбце
                                            min_thickness_column = min(list_thickness_column)
                                            # добавляем это значение в список минимальных значений столбцов
                                            list_min_thickness_column.append(min_thickness_column)
                                    # переменная всех значений толщин в столбце при поиске по номеру репорта
                                    if table_for_search_report:
                                        thickness_column = cur.execute(
                                            'SELECT {} from {}'.format(ii, table_for_search_report[0][i][0])).fetchall()
                                        # выбираем только вещественные значения
                                        for iii in thickness_column:
                                            # проверка если в столбце нет значений, то дальше, иначе...
                                            if not iii[0]:
                                                continue
                                            else:
                                                try:
                                                    if float(iii[0]):
                                                        list_thickness_column.append(float(iii[0]))
                                                except ValueError:
                                                    continue
                                        # если в столбце нет значений, то дальше, иначе...
                                        if not list_thickness_column:
                                            continue
                                        else:
                                            # минимальное значение толщины в столбце
                                            min_thickness_column = min(list_thickness_column)
                                            # добавляем это значение в список минимальных значений столбцов
                                            list_min_thickness_column.append(min_thickness_column)
                                    # переменная всех значений толщин в столбце при поиске по номеру work order
                                    if table_for_search_wo:
                                        thickness_column = cur.execute('SELECT {} from {}'.format(ii, table_for_search_wo[0][i][0])).fetchall()
                                        # выбираем только вещественные значения
                                        for iii in thickness_column:
                                            # проверка если в столбце нет значений, то дальше, иначе...
                                            if not iii[0]:
                                                continue
                                            else:
                                                try:
                                                    if float(iii[0]):
                                                        list_thickness_column.append(float(iii[0]))
                                                except ValueError:
                                                    continue
                                        # если в столбце нет значений, то дальше, иначе...
                                        if not list_thickness_column:
                                            continue
                                        else:
                                            # минимальное значение толщины в столбце
                                            min_thickness_column = min(list_thickness_column)
                                            # добавляем это значение в список минимальных значений столбцов
                                            list_min_thickness_column.append(min_thickness_column)
                                    # закрываем соединение с базой данных
                                    cur.close()
                            # после перебора всех допустимых столбцов выбираем минимальное значение global list_min_thickness
                            min_thickness = min(list_min_thickness_column)
                            list_min_thickness.append(min_thickness)
                            # высота одной таблицы tableView = количество строк в одной таблице * высоту одной строки +
                            # + высота строки названия столбцов
                            height = count_row_table_view[i] * one_row + one_row
                            # создаём переменную названия кнопок номеров репортов для вывода данных
                            if table_for_search_line:
                                button_for_table = table_for_search_line[i]
                                # переменная для поиска даты и work order репорта в таблице master
                                for_w_o = button_for_table[(button_for_table.index('_04') + 1):]
                                # подключаемся в базе данных
                                conn = sqlite3.connect(r'C:\Users\asus\PycharmProjects\YKR\YKR\DB\\' + db)
                                # conn = sqlite3.connect('reports_db.sqlite')
                                # conn.isolation_level = None
                                # cur = conn.cursor()
                                # подключаемся в базе данных
                                cur = conn.cursor()
                                # переменная номера work order
                                w_o = cur.execute('SELECT report_date, work_order FROM master WHERE report_number="{}"'.format(
                                    for_w_o)).fetchall()
                                # закрываем соединение
                                cur.close()
                                button_for_table = re.sub(r'_', '-', button_for_table)
                                ind = button_for_table.index('-04') + 1
                                # название кнопки по номеру репорта
                                second_underlining = button_for_table[ind:]
                                # добавляем к названию кнопки дату и work order
                                print(w_o)
                                second_underlining = second_underlining + '     Date: ' + w_o[0][0] + '     WO: ' + w_o[0][
                                    1] + '     min = ' + str(min_thickness) + '     UTT'
                            if table_for_search_drawing:
                                button_for_table = table_for_search_drawing[i]
                                # переменная для поиска даты и work order репорта в таблице master
                                for_w_o = button_for_table[(button_for_table.index('_04') + 1):]
                                # подключаемся в базе данных
                                conn = sqlite3.connect(r'C:\Users\asus\PycharmProjects\YKR\YKR\DB\\' + db)
                                # conn = sqlite3.connect('reports_db.sqlite')
                                # conn.isolation_level = None
                                # cur = conn.cursor()
                                # подключаемся в базе данных
                                cur = conn.cursor()
                                # переменная номера work order
                                w_o = cur.execute(
                                    'SELECT report_date, work_order FROM master WHERE report_number="{}"'.format(for_w_o)).fetchall()
                                # закрываем соединение
                                cur.close()
                                button_for_table = re.sub(r'_', '-', button_for_table)
                                ind = button_for_table.index('-04') + 1
                                # название кнопки по номеру репорта
                                second_underlining = button_for_table[ind:]
                                # добавляем к названию кнопки дату и work order
                                second_underlining = second_underlining + '     Date: ' + w_o[0][0] + '     WO: ' + w_o[0][
                                    1] + '     min = ' + str(min_thickness) + '     UTT'
                            # определяем глубину вложенности списка заданного для поиска репорта
                            if depthCount(table_for_search_report) == 1:
                                table_for_search_report = []
                            if table_for_search_report:
                                button_for_table = table_for_search_report[0][i][0]
                                # переменная для поиска даты и work order репорта в таблице master
                                for_w_o = button_for_table[(button_for_table.index('_04') + 1):]
                                # подключаемся в базе данных
                                conn = sqlite3.connect(r'C:\Users\asus\PycharmProjects\YKR\YKR\DB\\' + db)
                                # conn = sqlite3.connect('reports_db.sqlite')
                                # подключаемся в базе данных
                                cur = conn.cursor()
                                # переменная номера work order
                                w_o = cur.execute(
                                    'SELECT report_date, work_order FROM master WHERE report_number="{}"'.format(for_w_o)).fetchall()
                                # закрываем соединение
                                cur.close()
                                button_for_table = re.sub(r'_', '-', button_for_table)
                                ind = button_for_table.index('-04') + 1
                                # название кнопки по номеру репорта
                                second_underlining = button_for_table[ind:]
                                # добавляем к названию кнопки дату и work order
                                second_underlining = second_underlining + '     Date: ' + w_o[0][0] + '     WO: ' + w_o[0][
                                    1] + '     min = ' + str(min_thickness) + '     UTT'
                            # определяем глубину вложенности списка заданного для поиска work order
                            if depthCount(table_for_search_wo) == 1:
                                table_for_search_wo = []
                            if table_for_search_wo:
                                button_for_table = table_for_search_wo[0][i][0]
                                # номер work order из строки для поиска
                                w_o = line_for_search
                                # подключаемся в базе данных
                                conn = sqlite3.connect(r'C:\Users\asus\PycharmProjects\YKR\YKR\DB\\' + db)
                                # conn = sqlite3.connect('reports_db.sqlite')
                                # подключаемся в базе данных
                                cur = conn.cursor()
                                # переменная даты репорта
                                date_report = cur.execute('SELECT report_date FROM master WHERE work_order="{}"'.format(w_o)).fetchall()
                                # закрываем соединение
                                cur.close()
                                button_for_table = re.sub(r'_', '-', button_for_table)
                                ind = button_for_table.index('-04') + 1
                                # название кнопки по номеру репорта
                                second_underlining = button_for_table[ind:]
                                # добавляем к названию кнопки дату и work order
                                second_underlining = second_underlining + '     Date: ' + date_report[0][
                                    0] + '     WO: ' + w_o + '     min = ' + str(min_thickness) + '     UTT'

                            # задаём название кнопки по номеру репорта и помещаем внутрь frame
                            button_for_table = QPushButton(second_underlining, frame_for_table)
                            # координата отступа от левого края (меняется, когда происходит авторизация пользователя -
                            # появляется check box)
                            x1 = 20
                            # задаём размеры и место расположения кнопки во frame
                            button_for_table.setGeometry(QRect(x1, y1, 800, 20))
                            # задаём стиль шрифта
                            font_button_for_table = QFont()
                            font_button_for_table.setFamily(u"Calibri")
                            font_button_for_table.setPointSize(10)
                            button_for_table.setStyleSheet('text-align: left; font: bold italic')
                            button_for_table.setFont(font_button_for_table)
                            button_for_table.show()
                            # скрываем границы кнопки
                            button_for_table.setFlat(True)
                            # делаем кнопку переключателем
                            button_for_table.setCheckable(True)

                            # задаём флажок для каждой кнопки номера репорта
                            check_box[i] = QCheckBox(frame_for_table)
                            # задаём координаты флажка
                            check_box[i].move(0, y1)
                            if authorization == 0:
                                # скрываем флажок
                                check_box[i].hide()
                            elif authorization != 0:
                                check_box[i].show()

                            # задаём поле для вывода данных из базы данных, размещённую в области с полосой прокрутки
                            table_view[i] = QTableView(frame_for_table)
                            # устанавливаем координаты расположения таблиц в области с полосой прокрутки
                            list_button_for_table.append(button_for_table)
                            list_check_box.append(check_box[i])
                            list_table_view.append(table_view[i])
                            list_height_table_view.append(height)
                            table_view[i].show()
                            # сдвигаем все последующие кнопки и таблицы
                            y1 += 20
                            # создаём модель
                            sqm = QSqlQueryModel(parent=window)
                            # устанавливаем ширину столбцов под содержимое
                            table_view[i].horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
                            # устанавливаем высоту столбцов под содержимое
                            table_view[i].verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
                            # устанавливаем разный цвет фона для чётных и нечётных строк
                            table_view[i].setAlternatingRowColors(True)
                            table_view[i].setModel(sqm)
                            # создаём запрос и подключаемся в базе данных
                            # conn = sqlite3.connect('reports_db.sqlite')
                            # выводим данные в форму из найденных таблиц по номеру линии, чертежа или репорта
                            if len(table_for_search_line) > 0:
                                sqm.setQuery('SELECT * FROM {} WHERE Line LIKE "%{}%"'.format(table_for_search_line[i], line_for_search),
                                             db=QSqlDatabase(db))
                                             # db=QSqlDatabase('reports_db.sqlite'))
                            # выводим данные в форму из найденных таблиц по номеру чертежа в таблице
                            if len(table_for_search_drawing) > 0:
                                sqm.setQuery('SELECT * FROM {} WHERE Drawing LIKE "%{}%"'.format(table_for_search_drawing[i], line_for_search),
                                             db=QSqlDatabase(db))
                                             # db=QSqlDatabase('reports_db.sqlite'))
                            # выводим данные в форму из найденных таблиц по номеру репорта
                            if len(table_for_search_report) > 0:
                                if len(table_for_search_report[0]) > 0:
                                    sqm.setQuery('SELECT * FROM {}'.format(table_for_search_report[0][i][0]),
                                                 db=QSqlDatabase(db))
                                                 # db=QSqlDatabase('reports_db.sqlite'))
                            # выводим данные в форму из найденных таблиц по номеру word order
                            if len(table_for_search_wo) > 0:
                                if len(table_for_search_wo[0]) > 0:
                                    sqm.setQuery('SELECT * FROM {}'.format(table_for_search_wo[0][i][0]),
                                                 db=QSqlDatabase(db))
                                                 # db=QSqlDatabase('reports_db.sqlite'))

                            table_view[i].hide()
                            # обработка нажатия на кнопку с номером репорта в frame
                            button_for_table.clicked.connect(
                                lambda: visible_table_view(x1, list_table_view, list_button_for_table, list_check_box, list_height_table_view))
                            # активируем кнопку в левом верхнем углу таблицы для выделения всей таблицы
                            table_view[i].setCornerButtonEnabled(True)
                            # горизонтальная полоса прокрутки в пределах отображения одной таблицы
                            table_view[i].setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
                            # получаем номер столбца с номинальной толщиной
                            if 'Nominal_thickness' in name_column:
                                number_column_nominal_thickness = name_column.index('Nominal_thickness')

                                # переопределяем цвет для окрашивания столбца с номинальной толщиной
                                class ColorNominalThickness(QItemDelegate):
                                    def __init__(self):
                                        super().__init__()
                                        self.filter = ''

                                    def paint(self, painter, option, index):
                                        # выбираем зелёный цвет для столбца с номинальной толщиной
                                        painter.fillRect(option.rect, QColor(35, 198, 23, 180))
                                        return QItemDelegate.paint(self, painter, option, index)

                                # создаём модель
                                color_nominal_thickness = ColorNominalThickness()
                                # окрашиваем столбец с номинальной толщиной в зелёный цвет
                                table_view[i].setItemDelegateForColumn(number_column_nominal_thickness, color_nominal_thickness)
                            # добавляем найденную модель с данными в список для возможной дальнейшей распечатки
                            list_sqm.append(sqm)
                            name_column_for_print.append(name_column)
                            list_name_sheet_for_print.append(second_underlining)

                        scroll_area.show()
                        logger_with_user.info('Произведён поиск данных по номеру {}. Данные найдены'.format(line_search.text()))

                    # сообщение о том, что ничего не найдено
                    else:
                        QMessageBox.information(
                            window,
                            'Внимание',
                            'Ничего не найдено!'
                        )
                        logger_with_user.info('Произведён поиск данных по номеру {}. Данные НЕ найдены'.format(line_search.text()))
            # con.close()
    # сообщение об ошибке, если в поле для поиска ничего не введено
    else:
        QMessageBox.information(
            window,
            'Внимание',
            'Вы не ввели ни номер линии, ни номер чертежа, ни номер репорта, ни номер word order для поиска данных'
        )
        logger_with_user.info('Попытка поиска данных с пустой строкой поиска')


# список нажатых кнопок
list_button_for_table_true = []


# функция отображения и повторного скрытия таблиц в frame
# l_t_v = list_table_view = список всех таблиц
# l_b_t = list_button_for_table = список всех номеров репортов
# y_1 - координата строки с номером репорта
# y_2 = y_1 + 20 - координата таблицы (20 - высота строки с номером репорта)
# l_h_t_v = list_height_table_view = список высот таблиц (строка с названием колонок + все строки таблицы)
def visible_table_view(x1, l_t_v, l_b_t, l_ch_b, l_h_t_v):
    # y_1 - координата первой строки номера репорта
    y_1 = 0
    # y_2 - координата первой таблицы
    y_2 = 20
    ii = 0
    # список новых координат номеров репортов
    position_y1 = []
    # список новых координат таблиц
    position_y2 = []
    global list_button_for_table_true
    # обнуляем список нажатых кнопок
    list_button_for_table_true = []
    # список отжатых кнопок
    list_button_for_table_false = []
    for i in l_b_t:
        # если нажата
        if i.isChecked():
            list_button_for_table_true.append(ii)
        # если не нажата
        if not i.isChecked():
            list_button_for_table_false.append(ii)
        ii += 1
    # вычисляем новые координаты номеров репортов и таблиц в зависимости от списка нажатых (list_button_for_table_true)
    # и не нажатых (list_button_for_table_false) кнопок
    for i in range(len(l_h_t_v)):
        # если нажата кнопка номера репорта
        if list_button_for_table_true:
            # перебираем номера нажатых кнопок
            for ii in list_button_for_table_true:
                if i == ii:
                    # добавляем в список координату кнопки номера репорта
                    position_y1.append(y_1)
                    # меняем координату кнопки номера репорта, потому что она нажата и появляется таблица с данными
                    y_1 += 40 + l_h_t_v[i]
                    # добавляем в список координату таблицы с данными
                    position_y2.append(y_2)
                    # меняем координату кнопки номера репорта, потому что она нажата и появляется таблица с данными
                    y_2 += 40 + l_h_t_v[ii]
        # если НЕ нажата кнопка номера репорта
        if list_button_for_table_false:
            # перебираем номера НЕ нажатых кнопок
            for ii in list_button_for_table_false:
                if i == ii:
                    # добавляем в список Не нажатой координату кнопки номера репорта
                    position_y1.append(y_1)
                    y_1 += 20
                    # добавляем в список координату таблицы с данными при НЕ нажатой кнопки номера репорта
                    position_y2.append(y_2)
                    y_2 += 20
    # делаем таблицы видимыми или скрываем их в зависимости от статуса
    for b in list_button_for_table_true:
        l_b_t[b].move(x1, position_y1[b])
        l_t_v[b].setGeometry((QRect(0, position_y2[b], 1460, l_h_t_v[b])))
        # делаем таблицу из списка видимой
        l_t_v[b].setVisible(True)
        if l_ch_b[b]:
            # передвигаем флажок
            l_ch_b[b].move(0, position_y2[b] - 20)

    for bb in list_button_for_table_false:
        # передвигаем кнопку репорта
        l_b_t[bb].move(x1, position_y1[bb])
        # делаем таблицу из списка снова скрытой
        l_t_v[bb].hide()
        if l_ch_b[bb]:
            # передвигаем флажок
            l_ch_b[bb].move(0, position_y1[bb])


# нажатие на кнопку "Удалить"
def delete_report():
    # проверяем наличие областей tableView для вывода данных
    # если есть, то закрываем их, чтобы не наслаивались
    if window.findChildren(QTableView):
        open_tableview = window.findChildren(QTableView)
        for i in open_tableview:
            i.hide()
    # список статусов флажков напротив репортов (установлен или не установлен флажок)
    check_uncheck_report_for_delete = []
    # список порядковых номеров репортов для удаления
    list_index_for_delete = []
    # порядковый номер репортов для удаления
    index_report_for_delete = 0
    for i in list_check_box:
        try:
            check_uncheck_report_for_delete.append(i.checkState())
        except RuntimeError:
            continue

    # если статус "2", то номер репорта добавляем в список репортов на удаление
    for i in check_uncheck_report_for_delete:
        if i == 2:
            list_index_for_delete.append(index_report_for_delete)
        index_report_for_delete += 1
    # если выбран(ы) репорт для удаления, то
    if 2 in set(check_uncheck_report_for_delete):
        # спрашиваем, точно ли надо удалять репорт(ы)
        question_delete = QMessageBox()
        question_delete.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        question_delete.setWindowTitle('Внимание')
        question_delete.setText('Вы уверены, что хотите удалить данные репорты?')
        # если нажата кнопка "Да", то
        if question_delete.exec() == QMessageBox.Yes:
            # активатор отсутствия репорта в таблице master для обновления области вывода найденных данных после
            # удаления всех таблиц из репорта
            check_update_scroll_area = 0
            if list_index_for_delete:
                # то выбираем для удаления таблицу по номеру индекса (list_index_for_delete) в list_table_for_delete_report
                for i in list_index_for_delete:
                    if type(list_table_for_delete_report[0][i][0]) == str and type(list_table_for_delete_report[0][i]) == tuple:
                        # подключаемся в базе данных
                        conn = sqlite3.connect('reports_db.sqlite')
                        # conn.isolation_level = None
                        cur = conn.cursor()
                        cur.execute('DROP TABLE {}'.format(list_table_for_delete_report[0][i][0]))
                        conn.commit()
                        # обновляем данные в таблице master по удалённым таблицам update_master_by_delete(list_table_for_delete_report)
                        # если ни одного репорта нет в sqlite_master, то удаляем номер репорта из master
                        if not cur.execute('SELECT * FROM sqlite_master WHERE  name LIKE "%{}%"'.format(
                                list_table_for_delete_report[0][i][0][-15:])).fetchone():
                            cur.execute(
                                'DELETE from master WHERE report_number LIKE "%{}%"'.format(list_table_for_delete_report[0][i][0][-15:]))
                            conn.commit()
                            # cur.close()
                            check_update_scroll_area += 1
                        cur.close()
                        if check_update_scroll_area == 0:
                            # обновляем данные в таблице master по удалённым таблицам
                            update_master_by_delete(list_table_for_delete_report[0][i][0])

                        logger_with_user.warning('БЫЛА УДАЛЕНА ТАБЛИЦА ' + list_table_for_delete_report[0][i][0])
                    if type(list_table_for_delete_report[0][i][0]) == str and type(list_table_for_delete_report[0][i]) == str:
                        # подключаемся в базе данных
                        conn = sqlite3.connect('reports_db.sqlite')
                        # conn.isolation_level = None
                        cur = conn.cursor()
                        cur.execute('DROP TABLE {}'.format(list_table_for_delete_report[0][i]))
                        conn.commit()
                        # обновляем данные в таблице master по удалённым таблицам update_master_by_delete(list_table_for_delete_report)
                        # если ни одного репорта нет в sqlite_master, то удаляем номер репорта из master
                        if not cur.execute('SELECT * FROM sqlite_master WHERE  name LIKE "%{}%"'.format(
                                list_table_for_delete_report[0][i][-15:])).fetchone():
                            cur.execute(
                                'DELETE from master WHERE report_number LIKE "%{}%"'.format(list_table_for_delete_report[0][i][-15:]))
                            conn.commit()
                            # cur.close()
                            check_update_scroll_area += 1
                        cur.close()
                        if check_update_scroll_area == 0:
                            # обновляем данные в таблице master по удалённым таблицам
                            update_master_by_delete(list_table_for_delete_report[0][i])

                        logger_with_user.warning('БЫЛА УДАЛЕНА ТАБЛИЦА ' + list_table_for_delete_report[0][i])
            # cur.close()
            QMessageBox.information(window,
                                    'Внимание!',
                                    'Выбранные репорты удалены!')
            # Проверяем, остался ли такой репорт в master
            # если активатор изменён, т.е. > 0 - такого репорта больше нет в master, то скрываем область для вывода найденных данных
            if check_update_scroll_area > 0:
                scroll_area.hide()
            # Если активатор не изменён, т.е. = 0, то обновляем область для вывода найденных данных
            elif check_update_scroll_area == 0:
                search()


# функция обновления столбцов 'one_of' и 'list_table_report' таблицы master при удалении таблиц из неё
# l_t_f_d_r =  list_table_for_delete_report = список таблиц для удаления
def update_master_by_delete(l_t_f_d_r):
    # подключаемся в базе данных
    conn = sqlite3.connect('reports_db.sqlite')
    # conn.isolation_level = None
    cur = conn.cursor()
    # получаем номера всех записанных таблиц в виде строки
    variable_report_for_delete_from_master = cur.execute(
        'SELECT list_table_report FROM master WHERE list_table_report LIKE "%{}%"'.format(l_t_f_d_r)).fetchall()[0][0]
    # print(variable_report_for_delete_from_master)
    # определяем номер репорта для удаления
    variable = cur.execute('SELECT report_number FROM master WHERE list_table_report LIKE "%{}%"'.format(l_t_f_d_r)).fetchall()[0][0]
    # print(variable)
    # удаляем в найденной строке, с номерами всех записанных таблиц, выбранную таблицу
    variable_report_for_delete_from_master_new = variable_report_for_delete_from_master.replace('\'' + l_t_f_d_r + '\'', '')
    # print(variable_report_for_delete_from_master_new)
    # удаляем в найденной строке лишние символы новой строки
    variable_report_for_delete_from_master_new = variable_report_for_delete_from_master_new.replace('\n\n', '\n')

    # conn.commit()
    # получаем количество репортов в столбце 'one_of'
    # print(l_t_f_d_r)
    one_of_column = cur.execute(
        'SELECT one_of FROM master WHERE list_table_report LIKE "%{}%"'.format(l_t_f_d_r)).fetchall()[0][0]
    # print(one_of_column)
    # получаем номер позиции символа '/'
    index_one_of_load = one_of_column.index('/')
    # уменьшаем на 1 количество записанных таблиц в столбце 'one_of'
    one_of_load_new = str(int(one_of_column[:index_one_of_load]) - 1)
    # определяем новое значение 'one_of'
    one_of_new = one_of_load_new + str(one_of_column[index_one_of_load:])
    # обновляем ячейку с количеством записанных таблиц
    cur.execute('UPDATE master SET one_of = "{}" WHERE report_number = "{}"'.format(
        one_of_new, variable))
    # обновляем ячейку с номерами всех оставшихся таблиц

    conn.commit()
    cur.execute('UPDATE master SET list_table_report = "{}" WHERE report_number = "{}"'.format(
        variable_report_for_delete_from_master_new, variable))
    conn.commit()
    cur.close()
    # if cur.execute(
    #     'SELECT list_table_report FROM master WHERE list_table_report LIKE "%{}%"'.format(i[0])):
    #     # получаем номера всех записанных таблиц в виде строки
    #     variable_report_for_delete_from_master = cur.execute(
    #         'SELECT list_table_report FROM master WHERE list_table_report LIKE "%{}%"'.format(i[0])).fetchall()[0][0]
    #     # определяем номер репорта
    #     variable = cur.execute(
    #         'SELECT report_number FROM master WHERE list_table_report LIKE "%{}%"'.format(i[0])).fetchall()[0][0]
    #     # удаляем в найденной строке, с номерами всех записанных таблиц, выбранную таблицу
    #     variable_report_for_delete_from_master_new = variable_report_for_delete_from_master.replace(
    #         '\'' + i[0] + '\'', '')
    #     # удаляем в найденной строке лишние символы новой строки
    #     variable_report_for_delete_from_master_new = variable_report_for_delete_from_master_new.replace('\n\n', '\n')
    #     # обновляем ячейку с номерами всех оставшихся таблиц
    #     cur.execute('UPDATE master SET list_table_report = "{}" WHERE report_number = "{}"'.format(
    #         variable_report_for_delete_from_master_new, variable))
    #     conn.commit()

    # elif type(l_t_f_d_r) == list:
    # for i in l_t_f_d_r:
    #     if cur.execute(
    #             'SELECT list_table_report FROM master WHERE list_table_report LIKE "%{}%"'.format(i)):
    #         # получаем номера всех записанных таблиц в виде строки
    #         variable_report_for_delete_from_master = cur.execute(
    #             'SELECT list_table_report FROM master WHERE list_table_report LIKE "%{}%"'.format(i)).fetchall()[0][0]
    #         # определяем номер репорта
    #         variable = cur.execute(
    #             'SELECT report_number FROM master WHERE list_table_report LIKE "%{}%"'.format(i)).fetchall()[0][0]
    #         # удаляем в найденной строке, с номерами всех записанных таблиц, выбранную таблицу
    #         variable_report_for_delete_from_master_new = variable_report_for_delete_from_master.replace(
    #             '\'' + i + '\'', '')
    #         # удаляем в найденной строке лишние символы новой строки
    #         variable_report_for_delete_from_master_new = variable_report_for_delete_from_master_new.replace('\n\n', '\n')
    #         # обновляем ячейку с номерами всех оставшихся таблиц
    #         cur.execute('UPDATE master SET list_table_report = "{}" WHERE report_number = "{}"'.format(
    #             variable_report_for_delete_from_master_new, variable))
    #


# активатор статистики
check_statistic_master = 0


def statistic_master():
    # проверяем наличие областей tableView для вывода данных
    # если есть, то закрываем их, чтобы не наслаивались
    if window.findChildren(QTableView):
        open_tableview = window.findChildren(QTableView)
        for i in open_tableview:
            i.hide()
    # создаём соединение с базой данной
    con = QSqlDatabase.addDatabase('QSQLITE')
    # передаём имя базы данных для открытия
    con.setDatabaseName(r'C:\Users\asus\PycharmProjects\YKR\YKR\reports_db.sqlite')
    con.open()
    # если соединение не установлено, то сообщение об ошибке и выход
    if not con.open():
        logger_with_user.error('Отсутствует соединение с базой данных')
        sys.exit()
    else:
        # подключаемся в базе данных
        conn = sqlite3.connect('reports_db.sqlite')
        cur = conn.cursor()
        w = 100000
        frame_for_statistic = QFrame()
        # помещаем frame в область с полосой прокрутки
        scroll_area.setWidget(frame_for_statistic)
        # задаём размер frame
        frame_for_statistic.setGeometry(0, 0, 1460, w)
        frame_for_statistic.show()
        # задаём поле для вывода данных из базы данных, размещённую в области с полосой прокрутки
        statistic = QTableView(frame_for_statistic)
        # задаём размер области для отображения данных
        statistic.setGeometry(0, 0, 1460, w)
        # создаём модель
        sqmm = QSqlTableModel()
        sqmm.setTable('master')
        sqmm.select()
        # sqm = QSqlQueryModel(parent=window)
        # устанавливаем ширину столбцов под содержимое
        statistic.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        # устанавливаем высоту столбцов под содержимое
        statistic.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        # устанавливаем разный цвет фона для чётных и нечётных строк
        statistic.setAlternatingRowColors(True)
        # statistic.setModel(sqm)
        statistic.setModel(sqmm)
        # создаём запрос и сортируем по номеру репорта
        # sqm.setQuery('SELECT * FROM master ORDER BY report_number DESC', db=QSqlDatabase('reports_db.sqlite'))
        # активируем кнопку в левом верхнем углу таблицы для выделения всей таблицы
        # statistic.setCornerButtonEnabled(True)
        # горизонтальная полоса прокрутки в пределах отображения одной таблицы
        # statistic.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        # frame_for_statistic.setVisible(True)
        # statistic.setVisible(True)
        statistic.show()
        # scroll_area.show()
        cur.close()
        logger_with_user.info('Просмотр сводных данных из таблицы master')
    # закрываем соединение с базой данных
    con.close()
    # активатор, если вызвана статистика
    global check_statistic_master
    check_statistic_master = 1


# нажатие кнопки "Войти"
def log_in():
    # если ничего не введено в поля "Логин" и "Пароль"
    if line_login.text() == '' and line_password.text() == '':
        QMessageBox.information(
            window,
            'Внимание!',
            'Вы ничего не ввели в поля "Логин" и "Пароль"!!!',
            buttons=QMessageBox.Ok
        )
        logger_with_user.error('Попытка авторизоваться - не заполнены поля "Логин" и "Пароль"')
    # если ничего не введено в поле "Логин"
    elif line_login.text() == '':
        QMessageBox.information(
            window,
            'Внимание!',
            'Вы не заполнили поле "Логин"!!!',
            buttons=QMessageBox.Ok
        )
        logger_with_user.error(
            'Попытка авторизоваться - не заполнено поле "Логин", указан пароль - "{}"'.format(line_password.text()))
    # если ничего не введено в поле "Пароль"
    elif line_password.text() == '':
        QMessageBox.information(
            window,
            'Внимание!',
            'Вы не заполнили поле "Пароль"!!!',
            buttons=QMessageBox.Ok
        )
        logger_with_user.error(
            'Попытка авторизоваться - Не заполнено поле "Пароль", указан логин - "{}"'.format(line_login.text()))
    # если правильно введён логин и пароль
    elif line_login.text() == 'admin' and line_password.text() == 'admin':
        # делаем активными кнопки "Добавить", "Удалить", "Выйти", "Сводные данные"
        button_delete.setDisabled(False)
        button_log_out.setDisabled(False)
        button_add.setDisabled(False)
        button_statistic_master.setDisabled(False)
        # очищаем поле ввода логина и пароля
        line_login.clear()
        line_password.clear()
        # блокируем кнопку "Войти"
        button_log_in.setDisabled(True)
        logger_with_user.info('Пользователь авторизовался')
        # делаем видимые флажки
        if list_check_box:
            for i in list_check_box:
                i.show()
        # сигнал о том, что выполнена авторизация
        global authorization
        authorization += 1
    # если неправильно введён логин или пароль
    else:
        QMessageBox.information(
            window,
            'Внимание!',
            'Вы ввели не правильный логин или пароль!!!',
            buttons=QMessageBox.Ok
        )
        logger_with_user.error(
            'Попытка авторизоваться - Введён неверный логин "{}" или пароль "{}"'.format(line_login.text(), line_password.text()))


def log_out():
    # делаем НЕ активными кнопки "Добавить", "Редактировать", "Удалить", "Выйти", "Сводные данные", "Выйти"
    button_delete.setDisabled(True)
    button_log_out.setDisabled(True)
    button_add.setDisabled(True)
    button_statistic_master.setDisabled(True)
    # разблокируем кнопку "Войти"
    button_log_in.setDisabled(False)
    logger_with_user.info('Пользователь вышел')
    # сбрасываем на ноль авторизацию для отображения флажков
    global authorization
    authorization = 0
    # если перед выходом из авторизации показана НЕ статистика из master (check_statistic_master == 0),
    if check_statistic_master == 0:
        # то скрываем флажки
        if list_check_box:
            for i in list_check_box:
                i.hide()


# отображение области с gif анимацией при загрузке новых репортов
def gif_loading():
    # создаём объект label
    label_gif = QLabel()
    # помещаем label в область с полосой прокрутки
    scroll_area.setWidget(label_gif)
    # присваиваем уникальное объектное имя
    label_gif.setObjectName(u"Loading")
    # его размер
    label_gif.setGeometry(QRect(0, 0, 1465, 645))
    label_gif.setAlignment(Qt.AlignCenter)
    movie = QMovie(u"gif_loading.gif")
    label_gif.setMovie(movie)
    # запускаем gif
    movie.start()
    # отображаем gif
    scroll_area.show()
    # window.update()
    scroll_area.update()


# функция для вывода найденных открытых репортов и сводных данных на лист Excel для дальнейшей печати на принтер
def print_table():
    wbb = openpyxl.Workbook()
    # дата и время формирования файла Excel для печати
    date_time_for_print = datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")
    thin = Side(border_style="thin", color="000000")
    # если открыты "Сводные данные"
    if check_statistic_master == 1:
        # создаём новый лист на каждую таблицу
        sheet_for_print = wbb.create_sheet('Statistic')
        # вставляем в ячейку "A1" название столбца "Номер репорта"
        sheet_for_print.cell(row=1, column=1, value=str("Номер репорта"))
        # выделяем её жирным
        sheet_for_print.cell(row=1, column=1).font = Font(bold=True)
        # центрируем запись внутри
        sheet_for_print.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        # Устанавливаем ширину столбца "A"
        sheet_for_print.column_dimensions['A'].width = 30
        # выделяем её границами
        sheet_for_print.cell(row=1, column=1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        # вставляем в ячейку "B1" название столбца "Дата репорта"
        sheet_for_print.cell(row=1, column=2, value=str("Дата репорта"))
        # выделяем её жирным
        sheet_for_print.cell(row=1, column=2).font = Font(bold=True)
        # центрируем запись внутри
        sheet_for_print.cell(row=1, column=2).alignment = Alignment(horizontal='center', vertical='center')
        # Устанавливаем ширину столбца "B"
        sheet_for_print.column_dimensions['B'].width = 15
        # выделяем её границами
        sheet_for_print.cell(row=1, column=2).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        # вставляем в ячейку "C1" название столбца "Work order"
        sheet_for_print.cell(row=1, column=3, value=str("Work order"))
        # выделяем её жирным
        sheet_for_print.cell(row=1, column=3).font = Font(bold=True)
        # центрируем запись внутри
        sheet_for_print.cell(row=1, column=3).alignment = Alignment(horizontal='center', vertical='center')
        # Устанавливаем ширину столбца "C"
        sheet_for_print.column_dimensions['C'].width = 15
        # выделяем её границами
        sheet_for_print.cell(row=1, column=3).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        # вставляем в ячейку "D1" название столбца "Загружено таблиц в БД / всего таблиц в файле"
        sheet_for_print.cell(row=1, column=4, value=str("Загружено таблиц в БД / всего таблиц в файле"))
        # выделяем её жирным
        sheet_for_print.cell(row=1, column=4).font = Font(bold=True)
        # центрируем запись внутри и переносим по словам
        sheet_for_print.cell(row=1, column=4).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # Устанавливаем ширину столбца "D"
        sheet_for_print.column_dimensions['D'].width = 15
        # выделяем её границами
        sheet_for_print.cell(row=1, column=4).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        # вставляем в ячейку "E1" название столбца "Список таблиц в БД"
        sheet_for_print.cell(row=1, column=5, value=str("Список таблиц в БД"))
        # выделяем её жирным
        sheet_for_print.cell(row=1, column=5).font = Font(bold=True)
        # центрируем запись внутри
        sheet_for_print.cell(row=1, column=5).alignment = Alignment(horizontal='center', vertical='center')
        # Устанавливаем ширину столбца "E"
        sheet_for_print.column_dimensions['E'].width = 35
        # выделяем её границами
        sheet_for_print.cell(row=1, column=5).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        # закрепляем первую строку с названием столбцов
        sheet_for_print.freeze_panes = "A2"

        # подключаемся в базе данных
        conn = sqlite3.connect('reports_db.sqlite')
        cur = conn.cursor()
        # получаем все данные из таблицы master и сортируем по номеру репорта
        myself = cur.execute("SELECT * FROM master ORDER BY report_number DESC")
        # перебираем строки в master
        for i, row in enumerate(myself):
            # перебираем столбцы в master
            for j, value in enumerate(row):
                # записываем значения в ячейки Excel
                sheet_for_print.cell(row=i + 2, column=j + 1, value=row[j])
                # центрируем запись внутри
                sheet_for_print.cell(row=i + 2, column=j + 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                # выделяем её границами
                sheet_for_print.cell(row=i + 2, column=j + 1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        cur.close()
        # путь сохранения в папке с программой
        new_path_for_print_statistic = os.path.abspath(os.getcwd()) + '\\Statistic for print\\' + date_time_for_print[:7] + '\\'
        if not os.path.exists(new_path_for_print_statistic):
            # то создаём эту папку
            os.makedirs(new_path_for_print_statistic)
        # переменная имени файла с расширением для сохранения и последующего открытия
        name_for_print_statistic = str(date_time_for_print) + ' Statistic for print' + '.xlsx'
        # Удаление листа, создаваемого по умолчанию, при создании документа
        del wbb['Sheet']
        # сохраняем файл
        wbb.save(new_path_for_print_statistic + name_for_print_statistic)
        wbb.close()
        # и открываем его
        os.startfile(new_path_for_print_statistic + name_for_print_statistic)
        logger_with_user.info('Вывод на печать сводных данных\n' + new_path_for_print_statistic + name_for_print_statistic)

    # если открыты найденные репорты
    # перебираем все выборки данных из базы данных
    elif list_sqm:
        # если открыта хоть одна найденная таблица
        if list_button_for_table_true:
            # перебираем номера открытых репортов (выбранные для печати)
            for i in list_button_for_table_true:
                c = list_sqm[i]
                # индекс номера таблицы по порядку с '0'
                index_table_for_print = list_sqm.index(c)
                # создаём новый лист на каждую таблицу
                sheet_for_print = wbb.create_sheet(
                    str(list_name_sheet_for_print[index_table_for_print]).replace(':', '-')[:25])
                # вставляем в первую строку название кнопки по выбранной таблицу
                sheet_for_print.cell(row=1, column=1, value=str(list_name_sheet_for_print[index_table_for_print]))
                # выделяем её жирным
                sheet_for_print.cell(row=1, column=1).font = Font(bold=True)
                # объединяем в первой строке столбцы 'A:J'
                sheet_for_print.merge_cells('A1:J1')
                # вставляем во вторую строку названия столбцов
                for collll in range(len(name_column_for_print[index_table_for_print])):
                    sheet_for_print.cell(row=2, column=collll + 1,
                                         value=str(name_column_for_print[index_table_for_print][collll]))
                    # выделяем её жирным
                    sheet_for_print.cell(row=2, column=collll + 1).font = Font(bold=True)
                    # центрируем запись внутри
                    sheet_for_print.cell(row=2, column=collll + 1).alignment = Alignment(horizontal='center', vertical='center')
                    # закрепляем первую строку с названием кнопки, по которой выбрана таблица, и вторую с названиями столбцов
                    sheet_for_print.freeze_panes = "A3"
                    # выделяем её границами
                    sheet_for_print.cell(row=2, column=collll + 1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                ii = 2
                # проходим по всем строка выборки
                for row in range(c.rowCount()):
                    ii += 1
                    # обнуляем столбец с которого начинаем заполнять лист Excel
                    jj = 0
                    # по всем столбцам выборки
                    for column in range(c.columnCount()):
                        jj += 1
                        # получаем индекс строки и столбца в выборке по порядку
                        ind = c.index(row, column)
                        # заполняем лист Excel
                        sheet_for_print.cell(row=ii, column=jj, value=str(c.data(ind)))
                        # выделяем основные данные границами
                        sheet_for_print.cell(row=ii, column=jj).border = Border(top=thin, left=thin, right=thin, bottom=thin)
                # ручной автоподбор ширины столбцов по содержимому
                ascii_range = ['', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
                               'S', 'T', 'V', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI',
                               'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AV', 'AX', 'AY', 'AZ']
                # перебираем все заполненные столбцы
                for coll in range(1, jj + 1):
                    max_length_column = 0
                    # перебираем все заполненные строки
                    for roww in range(2, ii + 1):
                        sheet_for_print.cell(row=roww, column=int(index_nom_thickness_name_column) + 1).fill = PatternFill(
                            fgColor="77dd77",
                            fill_type="solid")
                        if len(str(sheet_for_print.cell(row=roww, column=coll).value)) > max_length_column:
                            max_length_column = len(str(sheet_for_print.cell(row=roww, column=coll).value))
                        # закрашиваем ячейки с минимальной толщиной
                        try:
                            if list_min_thickness[index_table_for_print] == float(sheet_for_print.cell(row=roww, column=coll).value):
                                sheet_for_print.cell(row=roww, column=coll).fill = PatternFill(fgColor="e34234", fill_type="solid")
                        except ValueError:
                            continue

                    # устанавливаем ширину заполненных столбцов по их содержимому
                    sheet_for_print.column_dimensions[ascii_range[coll]].width = max_length_column + 2

            # путь сохранения в папке с программой
            new_path_for_print = os.path.abspath(os.getcwd()) + '\\Report for print\\' + date_time_for_print[:7] + '\\'
            if not os.path.exists(new_path_for_print):
                # то создаём эту папку
                os.makedirs(new_path_for_print)
            # переменная имени файла с расширением для сохранения и последующего открытия
            name_for_print = str(date_time_for_print) + ' Report for print' + '.xlsx'
            # Удаление листа, создаваемого по умолчанию, при создании документа
            del wbb['Sheet']
            # сохраняем файл
            wbb.save(new_path_for_print + name_for_print)
            wbb.close()
            # и открываем его
            os.startfile(new_path_for_print + name_for_print)
            logger_with_user.info('Вывод на печать репорта(ов)\n' + new_path_for_print + name_for_print)


# нажатие кнопки "Войти"
button_log_in.clicked.connect(log_in)
# нажатие на кнопку Enter когда фокус (каретка - мигающий символ "|") находится в поле для ввода логина
line_login.returnPressed.connect(log_in)
# нажатие на кнопку Enter когда фокус (каретка - мигающий символ "|") находится в поле для ввода пароля
line_password.returnPressed.connect(log_in)

# нажатие на кнопку "Печать"
button_print.clicked.connect(print_table)

# нажатие на кнопку "Добавить"
button_add.clicked.connect(add_tables)

# нажатие на кнопку "Поиск"
button_search.clicked.connect(search)
# нажатие на кнопку Enter когда фокус (каретка - мигающий символ "|") находится в поле для ввода номера линии, чертежа,
# номера репорта или work order
line_search.returnPressed.connect(search)

# нажатие на кнопку "Удалить"
button_delete.clicked.connect(delete_report)

# нажатие на кнопку "Сводные данные"
button_statistic_master.clicked.connect(statistic_master)

# нажатие на кнопку "Выйти"
button_log_out.clicked.connect(log_out)


def main():
    try:
        window.show()

        sys.exit(app.exec_())
    finally:
        logger_with_user.info('Программа закрыта\n'
                              '--------------------------------------------------------------------------------')


if __name__ == '__main__':
    main()
